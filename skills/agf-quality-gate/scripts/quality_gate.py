# -*- coding: utf-8 -*-
"""
AGF Quality Gate — 策划文档质量门禁系统
用于在文档生成前后进行自动化质量校验

核心功能：
1. pre_check  — 生成前提取源文档指纹（规则数/表格数/数值/图片等）
2. post_check — 生成后扫描产出文档并与指纹对比
3. generate_report — 输出结构化质量报告（Markdown）

依赖: openpyxl（与 game-design-doc-template 共享）
"""

import os
import re
import json
from datetime import datetime
from collections import OrderedDict

try:
    from openpyxl import load_workbook
except ImportError:
    raise ImportError("需要 openpyxl 库。运行: pip install openpyxl")


# ==================== 配置 ====================

DEFAULT_OUTPUT_DIR = "G:/project_output"

# 质量规则默认阈值
DEFAULT_THRESHOLDS = {
    "rule_coverage_min": 0.90,        # 规则覆盖率最低 90%
    "table_coverage_min": 0.90,       # 表格覆盖率最低 90%
    "content_line_ratio_min": 0.70,   # 内容行比例最低 70%（SKILL.md 原有要求）
    "numeric_drift_max": 0.0,         # 数值偏差允许 0（严格模式）
    "format_score_min": 80,           # 格式评分最低 80/100
    "pending_items_flagged": True,    # 待决策项必须标注
}


# ==================== 指纹提取器 ====================

class DocumentFingerprint:
    """文档指纹——记录源文档或产出文档的关键结构化指标"""
    
    def __init__(self):
        self.total_rules = 0              # 规则总数（rule_title 级）
        self.total_content_lines = 0      # 正文内容行数（add_content 级）
        self.total_tables = 0             # 表格总数
        self.total_table_rows = 0         # 表格数据总行数
        self.total_images = 0             # 图片总数
        self.total_pending_items = 0      # 待决策项数
        self.sheet_count = 0              # 页签数量
        self.sheet_names = []             # 页签名称列表
        self.numeric_values = []          # 提取到的关键数值 [(位置描述, 值), ...]
        self.rule_titles = []             # 规则标题列表
        self.table_titles = []            # 表格标题列表
        self.section_titles = []          # 一级标题列表
        self.highlights = []              # 高亮/重点内容
        self.raw_detail = {}              # 逐页签详细数据
    
    def to_dict(self):
        return {
            "total_rules": self.total_rules,
            "total_content_lines": self.total_content_lines,
            "total_tables": self.total_tables,
            "total_table_rows": self.total_table_rows,
            "total_images": self.total_images,
            "total_pending_items": self.total_pending_items,
            "sheet_count": self.sheet_count,
            "sheet_names": self.sheet_names,
            "numeric_values_count": len(self.numeric_values),
            "rule_titles": self.rule_titles,
            "table_titles": self.table_titles,
            "section_titles": self.section_titles,
        }


def _extract_numbers(text):
    """从文本中提取数值（包括整数、小数、百分比）"""
    if not text or not isinstance(text, str):
        return []
    # 匹配: 100, 3.14, 50%, 0.5s, 1hour, x5, x100 等
    patterns = [
        r'(?<![.\w])\d+\.?\d*%',          # 百分比: 50%, 3.14%
        r'(?<![.\w])\d+\.?\d*(?:s|ms|min|hour|h|天|次|个|级|分)',  # 带单位数值
        r'[xX×]\d+',                        # 数量: x5, X100
        r'(?<![.\w])\d{2,}(?![.\w])',       # 2位以上纯数字（排除序号如"1."）
    ]
    results = []
    for pattern in patterns:
        matches = re.findall(pattern, text)
        results.extend(matches)
    return results


def _is_rule_title_cell(cell_value, col):
    """判断一个单元格是否是规则标题（C列，含"规则"字样或类似模式）"""
    if col != 3 or not cell_value or not isinstance(cell_value, str):
        return False
    text = cell_value.strip()
    # 匹配: "规则1：xxx", "规则N：xxx"
    if re.match(r'规则\d*[：:].+', text):
        return True
    return False


def _is_section_title(cell_value, col, font=None):
    """判断是否是一级标题（A列合并或B列大号加粗）"""
    if not cell_value or not isinstance(cell_value, str):
        return False
    text = cell_value.strip()
    # 【xxx】格式通常是一级标题
    if text.startswith('【') and text.endswith('】'):
        return True
    if font and font.size and font.size >= 14 and font.bold:
        return True
    return False


def _is_table_header_row(row_cells, start_col=2):
    """判断是否是表格表头行（连续多列有值且居中加粗）"""
    filled = 0
    for cell in row_cells:
        if cell.column >= start_col and cell.value:
            filled += 1
    return filled >= 3  # 至少3列有值


def _has_table_title_pattern(text):
    """判断是否是表格标题（如 "表1-xxx"）"""
    if not text or not isinstance(text, str):
        return False
    return bool(re.match(r'表\d+[-—]', text.strip()))


def extract_fingerprint_from_xlsx(file_path):
    """
    从 xlsx 文件提取文档指纹。
    
    适用于：
    1. 源文档（用户提供的原始策划文档）
    2. 产出文档（AI 生成的标准化文档）
    
    参数:
        file_path: xlsx 文件路径
    
    返回:
        DocumentFingerprint 对象
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")
    
    wb = load_workbook(file_path, read_only=True, data_only=True)
    fp = DocumentFingerprint()
    fp.sheet_count = len(wb.sheetnames)
    fp.sheet_names = list(wb.sheetnames)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_detail = {
            "content_lines": 0,
            "rules": [],
            "tables": [],
            "images": 0,
            "pending_items": 0,
            "numeric_values": [],
            "section_titles": [],
        }
        
        in_table = False
        table_row_count = 0
        current_table_title = None
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                # read_only 模式下可能出现 EmptyCell
                try:
                    val = cell.value
                    col = cell.column
                except AttributeError:
                    continue
                
                if not val or not isinstance(val, (str, int, float)):
                    continue
                
                text = str(val).strip()
                if not text:
                    continue
                
                # 检测一级标题
                if _is_section_title(text, col, getattr(cell, 'font', None)):
                    title_clean = text.strip('【】')
                    if title_clean not in [t for t in fp.section_titles]:
                        fp.section_titles.append(title_clean)
                        sheet_detail["section_titles"].append(title_clean)
                
                # 检测规则标题
                if _is_rule_title_cell(text, col):
                    fp.total_rules += 1
                    fp.rule_titles.append(text)
                    sheet_detail["rules"].append(text)
                
                # 检测表格标题
                if _has_table_title_pattern(text):
                    fp.total_tables += 1
                    fp.table_titles.append(text)
                    sheet_detail["tables"].append(text)
                    current_table_title = text
                    in_table = True
                    table_row_count = 0
                    continue
                
                # 表格行计数（简单启发：前一个标题是表格标题后的连续有值行）
                if in_table and col == 2 and text:
                    if _is_section_title(text, col) or _is_rule_title_cell(text, 3):
                        in_table = False
                    else:
                        table_row_count += 1
                        fp.total_table_rows += 1
                
                # 检测正文内容行（E列=5）
                if col == 5 and text:
                    fp.total_content_lines += 1
                    sheet_detail["content_lines"] += 1
                
                # 检测待决策项（含"待决策"字样）
                if '待决策' in text or '待确认' in text or '待讨论' in text:
                    fp.total_pending_items += 1
                    sheet_detail["pending_items"] += 1
                
                # 检测图片引用
                if '[嵌入' in text or '图片' in text.lower() or '.png' in text.lower() or '.jpg' in text.lower():
                    fp.total_images += 1
                    sheet_detail["images"] += 1
                
                # 提取数值
                nums = _extract_numbers(text)
                for n in nums:
                    location = f"{sheet_name}!{cell.coordinate}"
                    fp.numeric_values.append((location, n))
                    sheet_detail["numeric_values"].append((location, n))
        
        fp.raw_detail[sheet_name] = sheet_detail
    
    wb.close()
    return fp


def extract_fingerprint_from_text(text_content, source_name="文本输入"):
    """
    从纯文本/Markdown 内容提取指纹。
    
    适用于用户直接输入的文字描述、Markdown 文档等。
    
    参数:
        text_content: 文本内容字符串
        source_name: 来源名称
    
    返回:
        DocumentFingerprint 对象
    """
    fp = DocumentFingerprint()
    fp.sheet_count = 1
    fp.sheet_names = [source_name]
    
    lines = text_content.split('\n')
    for i, line in enumerate(lines):
        text = line.strip()
        if not text:
            continue
        
        fp.total_content_lines += 1
        
        # 检测标题（Markdown # 或 【xxx】）
        if text.startswith('#') or (text.startswith('【') and text.endswith('】')):
            title = text.lstrip('#').strip().strip('【】')
            fp.section_titles.append(title)
        
        # 检测规则模式
        if re.match(r'规则\d*[：:].+', text):
            fp.total_rules += 1
            fp.rule_titles.append(text)
        
        # 检测表格标题
        if _has_table_title_pattern(text):
            fp.total_tables += 1
            fp.table_titles.append(text)
        
        # 检测待决策
        if '待决策' in text or '待确认' in text:
            fp.total_pending_items += 1
        
        # 提取数值
        nums = _extract_numbers(text)
        for n in nums:
            fp.numeric_values.append((f"line:{i+1}", n))
    
    return fp


# ==================== 质量对比引擎 ====================

class QualityCheckResult:
    """单项检查结果"""
    
    def __init__(self, name, passed, score, actual, expected, detail=""):
        self.name = name
        self.passed = passed
        self.score = score        # 0-100
        self.actual = actual
        self.expected = expected
        self.detail = detail
    
    def to_dict(self):
        return {
            "name": self.name,
            "passed": self.passed,
            "score": self.score,
            "actual": self.actual,
            "expected": self.expected,
            "detail": self.detail,
        }


def compare_fingerprints(source_fp, output_fp, thresholds=None):
    """
    对比源文档指纹与产出文档指纹，生成质量检查结果列表。
    
    参数:
        source_fp: 源文档指纹 (DocumentFingerprint)
        output_fp: 产出文档指纹 (DocumentFingerprint)
        thresholds: 质量阈值配置 (dict)，为 None 时使用默认值
    
    返回:
        list[QualityCheckResult]
    """
    if thresholds is None:
        thresholds = DEFAULT_THRESHOLDS
    
    results = []
    
    # 1. 规则覆盖率
    if source_fp.total_rules > 0:
        coverage = output_fp.total_rules / source_fp.total_rules
        min_req = thresholds.get("rule_coverage_min", 0.90)
        
        # 查找丢失的规则
        source_rules_normalized = [r.strip() for r in source_fp.rule_titles]
        output_rules_normalized = [r.strip() for r in output_fp.rule_titles]
        missing = [r for r in source_rules_normalized if r not in output_rules_normalized]
        
        detail = ""
        if missing:
            detail = f"可能丢失的规则: {', '.join(missing[:5])}"
            if len(missing) > 5:
                detail += f" ...等共{len(missing)}条"
        
        results.append(QualityCheckResult(
            name="规则覆盖率",
            passed=coverage >= min_req,
            score=min(100, int(coverage * 100)),
            actual=f"{output_fp.total_rules}/{source_fp.total_rules} ({coverage:.0%})",
            expected=f"≥ {min_req:.0%}",
            detail=detail,
        ))
    else:
        results.append(QualityCheckResult(
            name="规则覆盖率",
            passed=True,
            score=100,
            actual="源文档无规则定义",
            expected="N/A",
            detail="源文档为纯文本/概述结构，跳过规则覆盖检查",
        ))
    
    # 2. 内容行覆盖率
    if source_fp.total_content_lines > 0:
        ratio = output_fp.total_content_lines / source_fp.total_content_lines
        min_req = thresholds.get("content_line_ratio_min", 0.70)
        results.append(QualityCheckResult(
            name="内容行覆盖率",
            passed=ratio >= min_req,
            score=min(100, int(ratio * 100)),
            actual=f"{output_fp.total_content_lines}/{source_fp.total_content_lines} ({ratio:.0%})",
            expected=f"≥ {min_req:.0%}",
            detail="产出文档的 add_content 调用次数 vs 源文档有效内容行数",
        ))
    
    # 3. 表格覆盖率
    if source_fp.total_tables > 0:
        coverage = output_fp.total_tables / source_fp.total_tables
        min_req = thresholds.get("table_coverage_min", 0.90)
        
        missing_tables = [t for t in source_fp.table_titles if t not in output_fp.table_titles]
        detail = ""
        if missing_tables:
            detail = f"可能丢失的表格: {', '.join(missing_tables[:5])}"
        
        results.append(QualityCheckResult(
            name="表格覆盖率",
            passed=coverage >= min_req,
            score=min(100, int(coverage * 100)),
            actual=f"{output_fp.total_tables}/{source_fp.total_tables} ({coverage:.0%})",
            expected=f"≥ {min_req:.0%}",
            detail=detail,
        ))
    
    # 4. 数值一致性
    source_nums = set(n for _, n in source_fp.numeric_values)
    output_nums = set(n for _, n in output_fp.numeric_values)
    if source_nums:
        preserved = source_nums & output_nums
        missing_nums = source_nums - output_nums
        preservation_rate = len(preserved) / len(source_nums)
        
        detail = ""
        if missing_nums:
            sample = list(missing_nums)[:10]
            detail = f"未在产出中找到的数值: {', '.join(sample)}"
            if len(missing_nums) > 10:
                detail += f" ...等共{len(missing_nums)}个"
        
        results.append(QualityCheckResult(
            name="数值一致性",
            passed=preservation_rate >= 0.80,
            score=min(100, int(preservation_rate * 100)),
            actual=f"{len(preserved)}/{len(source_nums)} ({preservation_rate:.0%})",
            expected="≥ 80%",
            detail=detail,
        ))
    
    # 5. 一级标题覆盖
    if source_fp.section_titles:
        source_sections = set(source_fp.section_titles)
        output_sections = set(output_fp.section_titles)
        covered = source_sections & output_sections
        missing_sections = source_sections - output_sections
        coverage = len(covered) / len(source_sections) if source_sections else 1.0
        
        detail = ""
        if missing_sections:
            detail = f"丢失的标题模块: {', '.join(missing_sections)}"
        
        results.append(QualityCheckResult(
            name="模块标题覆盖",
            passed=coverage >= 0.90,
            score=min(100, int(coverage * 100)),
            actual=f"{len(covered)}/{len(source_sections)} ({coverage:.0%})",
            expected="≥ 90%",
            detail=detail,
        ))
    
    # 6. 待决策项标注检查
    if source_fp.total_pending_items > 0:
        results.append(QualityCheckResult(
            name="待决策项保留",
            passed=output_fp.total_pending_items >= source_fp.total_pending_items,
            score=100 if output_fp.total_pending_items >= source_fp.total_pending_items else 50,
            actual=f"{output_fp.total_pending_items}个",
            expected=f"≥ {source_fp.total_pending_items}个",
            detail="源文档中的待决策/待确认项必须在产出中保留",
        ))
    
    # 7. 格式规范检查（仅针对标准化产出文档）
    format_score = _check_format_compliance(output_fp)
    results.append(QualityCheckResult(
        name="格式规范",
        passed=format_score >= thresholds.get("format_score_min", 80),
        score=format_score,
        actual=f"{format_score}/100",
        expected=f"≥ {thresholds.get('format_score_min', 80)}/100",
        detail="检查页签结构、标题层级、表格格式",
    ))
    
    return results


def _check_format_compliance(fp):
    """
    检查产出文档的格式规范性。
    评分项：
    - 4个标准页签存在 (+25)
    - 页签1有文档信息 (+15)
    - 页签2有设计内容 (+25)
    - 页签3有数值表格 (+15)
    - 页签4有打点设计 (+10)
    - 有一级标题结构  (+10)
    """
    score = 0
    
    # 4个标准页签
    standard_sheets = {'文档信息', '设计内容', '数值表格设计', 'tlog及打点设计'}
    existing = set(fp.sheet_names)
    matched = standard_sheets & existing
    score += int(len(matched) / 4 * 25)
    
    # 页签1: 文档信息
    if '文档信息' in fp.raw_detail:
        detail = fp.raw_detail['文档信息']
        if detail.get('content_lines', 0) > 0 or detail.get('section_titles'):
            score += 15
    
    # 页签2: 设计内容
    if '设计内容' in fp.raw_detail:
        detail = fp.raw_detail['设计内容']
        if detail.get('rules') or detail.get('content_lines', 0) > 5:
            score += 25
    
    # 页签3: 数值表格
    if '数值表格设计' in fp.raw_detail:
        detail = fp.raw_detail['数值表格设计']
        if detail.get('tables'):
            score += 15
    
    # 页签4: 打点设计
    if 'tlog及打点设计' in fp.raw_detail:
        detail = fp.raw_detail['tlog及打点设计']
        if detail.get('tables'):
            score += 10
    
    # 一级标题结构
    if len(fp.section_titles) >= 2:
        score += 10
    
    return min(100, score)


# ==================== 质量报告生成器 ====================

def generate_quality_report(source_fp, output_fp, check_results,
                            source_name="源文档", output_name="产出文档",
                            output_dir=None):
    """
    生成 Markdown 格式的质量报告。
    
    参数:
        source_fp: 源文档指纹
        output_fp: 产出文档指纹
        check_results: 质量检查结果列表
        source_name: 源文档名称
        output_name: 产出文档名称
        output_dir: 报告输出目录（None 则不自动保存，只返回文本）
    
    返回:
        str: Markdown 格式的报告文本
        str: 报告文件路径（如果 output_dir 不为 None）
    """
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    # 计算总分
    total_score = sum(r.score for r in check_results) / len(check_results) if check_results else 0
    all_passed = all(r.passed for r in check_results)
    
    # 等级判定
    if total_score >= 95:
        grade = "🟢 优秀 (A)"
    elif total_score >= 85:
        grade = "🟢 良好 (B)"
    elif total_score >= 70:
        grade = "🟡 合格 (C)"
    elif total_score >= 50:
        grade = "🟠 需改进 (D)"
    else:
        grade = "🔴 不合格 (F)"
    
    # 构建报告
    lines = []
    lines.append(f"# 📋 策划文档质量报告")
    lines.append("")
    lines.append(f"> ⏰ 检查时间：{now}")
    lines.append(f"> 📄 源文档：{source_name}")
    lines.append(f"> 📝 产出文档：{output_name}")
    lines.append(f"> 🎯 综合评分：**{total_score:.0f}/100** | 等级：**{grade}**")
    lines.append(f"> {'✅ 全部通过' if all_passed else '⚠️ 存在未通过项'}")
    lines.append("")
    
    # 总览表
    lines.append("## 一、检查总览")
    lines.append("")
    lines.append("| # | 检查项 | 结果 | 得分 | 实际值 | 要求 |")
    lines.append("|---|--------|------|------|--------|------|")
    for i, r in enumerate(check_results, 1):
        icon = "✅" if r.passed else "❌"
        lines.append(f"| {i} | {r.name} | {icon} | {r.score}/100 | {r.actual} | {r.expected} |")
    lines.append("")
    
    # 详细分析
    failed = [r for r in check_results if not r.passed]
    if failed:
        lines.append("## 二、⚠️ 未通过项详情")
        lines.append("")
        for r in failed:
            lines.append(f"### ❌ {r.name}")
            lines.append(f"- **实际值**: {r.actual}")
            lines.append(f"- **要求**: {r.expected}")
            if r.detail:
                lines.append(f"- **详情**: {r.detail}")
            lines.append("")
    
    # 指纹对比
    lines.append("## 三、文档指纹对比")
    lines.append("")
    lines.append("| 指标 | 源文档 | 产出文档 | 差异 |")
    lines.append("|------|--------|----------|------|")
    
    metrics = [
        ("一级标题数", len(source_fp.section_titles), len(output_fp.section_titles)),
        ("规则标题数", source_fp.total_rules, output_fp.total_rules),
        ("正文内容行数", source_fp.total_content_lines, output_fp.total_content_lines),
        ("表格数", source_fp.total_tables, output_fp.total_tables),
        ("表格数据行数", source_fp.total_table_rows, output_fp.total_table_rows),
        ("待决策项数", source_fp.total_pending_items, output_fp.total_pending_items),
        ("关键数值数", len(source_fp.numeric_values), len(output_fp.numeric_values)),
        ("页签数", source_fp.sheet_count, output_fp.sheet_count),
    ]
    
    for name, src, out in metrics:
        diff = out - src
        diff_str = f"+{diff}" if diff > 0 else str(diff) if diff < 0 else "="
        diff_icon = "🟢" if diff >= 0 else "🔴"
        lines.append(f"| {name} | {src} | {out} | {diff_icon} {diff_str} |")
    lines.append("")
    
    # 规则标题清单
    if source_fp.rule_titles or output_fp.rule_titles:
        lines.append("## 四、规则标题清单")
        lines.append("")
        lines.append("### 源文档规则")
        for i, t in enumerate(source_fp.rule_titles, 1):
            in_output = "✅" if t in output_fp.rule_titles else "❌ 丢失"
            lines.append(f"{i}. {t} — {in_output}")
        lines.append("")
        
        # 产出中新增的规则
        new_rules = [t for t in output_fp.rule_titles if t not in source_fp.rule_titles]
        if new_rules:
            lines.append("### 产出新增规则")
            for t in new_rules:
                lines.append(f"- 🆕 {t}")
            lines.append("")
    
    # 改进建议
    lines.append("## 五、改进建议")
    lines.append("")
    if all_passed:
        lines.append("✅ 所有检查项均通过，文档质量良好。")
    else:
        for r in failed:
            if r.name == "规则覆盖率":
                lines.append("- 🔧 **补充丢失的规则**：逐条检查源文档中的规则标题，确保每条都有对应的 `add_rule_title` 调用")
            elif r.name == "内容行覆盖率":
                lines.append("- 🔧 **补充内容行**：源文档中每个条目/说明都应对应一个 `add_content` 调用，避免概括性省略")
            elif r.name == "表格覆盖率":
                lines.append("- 🔧 **补充丢失的表格**：检查页签3是否遗漏了源文档中的配置表")
            elif r.name == "数值一致性":
                lines.append("- 🔧 **核实数值**：部分源文档中的关键数值未在产出中找到，需逐一确认")
            elif r.name == "格式规范":
                lines.append("- 🔧 **修正格式**：确保4个标准页签完整，层级缩进正确")
    lines.append("")
    
    # 签名
    lines.append("---")
    lines.append(f"*AGF Quality Gate v1.0 | 自动生成于 {now}*")
    
    report_text = "\n".join(lines)
    
    # 保存文件
    report_path = None
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        safe_name = output_name.replace('.xlsx', '').replace('/', '_').replace('\\', '_')
        report_filename = f"{safe_name}_质量报告.md"
        report_path = os.path.join(output_dir, report_filename)
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(report_text)
    
    return report_text, report_path


# ==================== 一站式 API ====================

def quality_check(source_path, output_path, thresholds=None, 
                  save_report=True, report_dir=None):
    """
    一站式质量检查：提取指纹 → 对比 → 生成报告。
    
    这是外部调用的主入口。
    
    参数:
        source_path: 源文档路径（xlsx 或包含文本的路径）
        output_path: 产出文档路径（xlsx）
        thresholds: 质量阈值（可选）
        save_report: 是否保存报告文件
        report_dir: 报告保存目录（默认与产出文档同目录）
    
    返回:
        dict: {
            "passed": bool,
            "score": float,
            "grade": str,
            "report_text": str,
            "report_path": str or None,
            "source_fingerprint": dict,
            "output_fingerprint": dict,
            "check_results": list[dict],
        }
    """
    # 提取指纹
    if source_path.endswith('.xlsx') or source_path.endswith('.xls'):
        source_fp = extract_fingerprint_from_xlsx(source_path)
    elif os.path.isfile(source_path):
        with open(source_path, 'r', encoding='utf-8') as f:
            text = f.read()
        source_fp = extract_fingerprint_from_text(text, os.path.basename(source_path))
    else:
        source_fp = extract_fingerprint_from_text(source_path, "文本输入")
    
    output_fp = extract_fingerprint_from_xlsx(output_path)
    
    # 对比
    check_results = compare_fingerprints(source_fp, output_fp, thresholds)
    
    # 报告
    source_name = os.path.basename(source_path) if os.path.isfile(source_path) else "文本输入"
    output_name = os.path.basename(output_path)
    
    if report_dir is None:
        report_dir = os.path.dirname(output_path) if save_report else None
    
    report_text, report_path = generate_quality_report(
        source_fp, output_fp, check_results,
        source_name=source_name,
        output_name=output_name,
        output_dir=report_dir if save_report else None,
    )
    
    # 汇总
    total_score = sum(r.score for r in check_results) / len(check_results) if check_results else 0
    all_passed = all(r.passed for r in check_results)
    
    if total_score >= 95:
        grade = "A"
    elif total_score >= 85:
        grade = "B"
    elif total_score >= 70:
        grade = "C"
    elif total_score >= 50:
        grade = "D"
    else:
        grade = "F"
    
    return {
        "passed": all_passed,
        "score": round(total_score, 1),
        "grade": grade,
        "report_text": report_text,
        "report_path": report_path,
        "source_fingerprint": source_fp.to_dict(),
        "output_fingerprint": output_fp.to_dict(),
        "check_results": [r.to_dict() for r in check_results],
    }


# ==================== CLI 入口 ====================

if __name__ == '__main__':
    import sys
    
    if len(sys.argv) < 3:
        print("用法: python quality_gate.py <源文档路径> <产出文档路径> [报告输出目录]")
        print("示例: python quality_gate.py source.xlsx output.xlsx G:/project_output")
        sys.exit(1)
    
    source = sys.argv[1]
    output = sys.argv[2]
    report_dir = sys.argv[3] if len(sys.argv) > 3 else None
    
    result = quality_check(source, output, report_dir=report_dir)
    
    print(f"\n{'='*60}")
    print(f"Quality Gate - Check Complete")
    print(f"{'='*60}")
    print(f"Score: {result['score']}/100 (Grade: {result['grade']})")
    print(f"Passed: {'YES' if result['passed'] else 'NO'}")
    if result['report_path']:
        print(f"Report: {result['report_path']}")
    print(f"{'='*60}\n")
    
    # 打印详细结果
    for r in result['check_results']:
        icon = "[OK]" if r['passed'] else "[FAIL]"
        print(f"  {icon} {r['name']}: {r['actual']} (req: {r['expected']})")

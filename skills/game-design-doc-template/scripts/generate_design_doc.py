# -*- coding: utf-8 -*-
"""
策划文档标准版生成脚本
用于生成标准化的Excel策划文档

依赖: openpyxl
安装: pip install openpyxl

层级缩进规范（列号对应关系）：
  A列(1): 留空
  B列(2): 一级标题(整行合并) / 二级标题 / 三级标题 / 单行内容 / 表格标题
  C列(3): 规则标题(rule_title)
  D列(4): 标签(label) / 待决策项(pending_item)
  E列(5): 内容(content)

颜色层级渐变：
  海军蓝(#2F5496) → 标准蓝(#4472C4) → 钢蓝(#8FAADC) → 天蓝(#B4C6E7) → 浅蓝灰(#D6DCE4)
  一级标题          二级标题            三级标题          规则标题          标签
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
import os

# ==================== 全局计数器 ====================
_auto_number_counter = 0

def reset_auto_number():
    """重置自动编号计数器。在每个规则的"1、规则说明"标签之前调用。"""
    global _auto_number_counter
    _auto_number_counter = 0

def get_next_number():
    """获取下一个递增序号"""
    global _auto_number_counter
    _auto_number_counter += 1
    return _auto_number_counter

# ==================== 颜色常量 ====================
NAVY       = '2F5496'   # 一级标题底色 - 深海军蓝
BLUE       = '4472C4'   # 二级标题底色 / 表头底色 - 标准蓝
STEEL      = '8FAADC'   # 三级标题底色 - 钢蓝
SKY        = 'B4C6E7'   # 规则标题底色 - 天蓝
LIGHT_BLUE = 'D6DCE4'   # 标签底色 - 浅蓝灰
PALE_BLUE  = 'D9E2F3'   # 表格标题底色 / 文档信息标签底色 - 淡蓝
LIGHT_GRAY = 'F2F2F2'   # 表格交替行底色
WARM_YELLOW= 'FFF2CC'   # 待决策项底色
WHITE      = 'FFFFFF'
DARK_GRAY  = '333333'   # 正文字色 - 深灰（比纯黑更舒适）
MID_GRAY   = '808080'   # 注释字色
LINK_BLUE  = '0563C1'   # 链接字色
RED        = 'C00000'   # 重点/待决策字色

# ==================== 样式定义 ====================

# 字体
font_title_1     = Font(name='微软雅黑', size=16, bold=True, color=WHITE)
font_title_2     = Font(name='微软雅黑', size=13, bold=True, color=WHITE)
font_title_3     = Font(name='微软雅黑', size=12, bold=True, color=NAVY)
font_rule_title  = Font(name='微软雅黑', size=11, bold=True, color=NAVY)
font_label       = Font(name='微软雅黑', size=11, bold=True, color=DARK_GRAY)
font_content     = Font(name='微软雅黑', size=11, bold=False, color=DARK_GRAY)
font_comment     = Font(name='微软雅黑', size=11, bold=False, color=MID_GRAY)
font_link        = Font(name='微软雅黑', size=11, bold=False, color=LINK_BLUE)
font_highlight   = Font(name='微软雅黑', size=11, bold=True, color=RED)
font_table_header= Font(name='微软雅黑', size=11, bold=True, color=WHITE)
font_table_content=Font(name='微软雅黑', size=11, bold=False, color=DARK_GRAY)
font_doc_label   = Font(name='微软雅黑', size=11, bold=True, color=NAVY)
font_doc_value   = Font(name='微软雅黑', size=11, bold=False, color=DARK_GRAY)

# 填充色
fill_title_1     = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
fill_title_2     = PatternFill(start_color=BLUE, end_color=BLUE, fill_type='solid')
fill_title_3     = PatternFill(start_color=STEEL, end_color=STEEL, fill_type='solid')
fill_rule_title  = PatternFill(start_color=SKY, end_color=SKY, fill_type='solid')
fill_label       = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
fill_table_title = PatternFill(start_color=PALE_BLUE, end_color=PALE_BLUE, fill_type='solid')
fill_table_header= PatternFill(start_color=BLUE, end_color=BLUE, fill_type='solid')
fill_table_alt   = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
fill_pending     = PatternFill(start_color=WARM_YELLOW, end_color=WARM_YELLOW, fill_type='solid')
fill_doc_label   = PatternFill(start_color=PALE_BLUE, end_color=PALE_BLUE, fill_type='solid')

# 边框
thin_border = Border(
    left=Side(style='thin', color='C0C0C0'),
    right=Side(style='thin', color='C0C0C0'),
    top=Side(style='thin', color='C0C0C0'),
    bottom=Side(style='thin', color='C0C0C0')
)

# 对齐
align_left   = Alignment(horizontal='left', vertical='center', wrap_text=False)
align_center = Alignment(horizontal='center', vertical='center', wrap_text=False)


# ==================== 基础函数 ====================

def create_workbook():
    """
    创建工作簿和4个标准页签。
    返回: (wb, ws1, ws2, ws3, ws4)
    """
    wb = Workbook()
    ws1 = wb.active
    ws1.title = '文档信息'
    ws2 = wb.create_sheet('设计内容')
    ws3 = wb.create_sheet('数值表格设计')
    ws4 = wb.create_sheet('tlog及打点设计')
    return wb, ws1, ws2, ws3, ws4


def setup_column_widths(ws):
    """设置标准列宽"""
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 65
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20


# ==================== 标题函数 ====================

def add_title_1(ws, row, text):
    """一级标题 - 合并A~E列, 海军蓝底白字, 16pt"""
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = font_title_1
    cell.fill = fill_title_1
    cell.alignment = align_left
    ws.row_dimensions[row].height = 36
    return row + 1


def add_title_2(ws, row, text):
    """二级标题 - B列, 标准蓝底白字, 13pt"""
    cell = ws.cell(row=row, column=2, value=text)
    cell.font = font_title_2
    cell.fill = fill_title_2
    cell.alignment = align_left
    ws.row_dimensions[row].height = 30
    return row + 1


def add_title_3(ws, row, text):
    """三级标题 - B列, 钢蓝底深蓝字, 12pt"""
    cell = ws.cell(row=row, column=2, value=text)
    cell.font = font_title_3
    cell.fill = fill_title_3
    cell.alignment = align_left
    ws.row_dimensions[row].height = 26
    return row + 1


# ==================== 规则相关函数 ====================

def add_rule_title(ws, row, text):
    """规则标题 - C列, 天蓝底深蓝字加粗11pt"""
    cell = ws.cell(row=row, column=3, value=text)
    cell.font = font_rule_title
    cell.fill = fill_rule_title
    cell.alignment = align_left
    ws.row_dimensions[row].height = 24
    return row + 1


def add_label(ws, row, text):
    """标签 - D列, 浅蓝灰底深灰字加粗11pt。文本需带序号如「1、规则说明」"""
    cell = ws.cell(row=row, column=4, value=text)
    cell.font = font_label
    cell.fill = fill_label
    cell.alignment = align_left
    ws.row_dimensions[row].height = 22
    return row + 1


def add_content(ws, row, text, is_comment=False, is_link=False, is_highlight=False, auto_number=False, is_sub=False):
    """
    内容 - E列, 11pt。

    参数:
        auto_number: True时使用全局计数器递增序号(1. 2. 3...)。
                     每个规则开始前需调用 reset_auto_number()。
        is_sub: True时表示子层级内容，会在文本前添加"  · "缩进前缀。
                用于表达规则说明中的嵌套子条目。
    """
    if is_sub and text:
        text = f"  · {text}"
    elif auto_number and text and not text.strip()[0].isdigit():
        num = get_next_number()
        text = f"{num}. {text}"

    cell = ws.cell(row=row, column=5, value=text)
    if is_comment:
        cell.font = font_comment
    elif is_link:
        cell.font = font_link
    elif is_highlight:
        cell.font = font_highlight
    elif is_sub:
        cell.font = Font(name='微软雅黑', size=11, bold=False, color=DARK_GRAY)
    else:
        cell.font = font_content
    cell.alignment = align_left
    ws.row_dimensions[row].height = 22
    return row + 1


def add_single_line(ws, row, text):
    """单行内容 - B列, 正文样式11pt"""
    cell = ws.cell(row=row, column=2, value=text)
    cell.font = font_content
    cell.alignment = align_left
    ws.row_dimensions[row].height = 22
    return row + 1


def add_pending_item(ws, row, text):
    """待决策项 - D列, 暖黄底红字加粗"""
    cell = ws.cell(row=row, column=4, value=text)
    cell.font = font_highlight
    cell.fill = fill_pending
    cell.alignment = align_left
    ws.row_dimensions[row].height = 24
    return row + 1


# ==================== 页签1专用：文档信息 ====================

def add_doc_info(ws, row, label, value):
    """
    文档信息行（页签1）。
    B列=标签(淡蓝底深蓝字居中), C~E列合并=内容。
    """
    cell_label = ws.cell(row=row, column=2, value=label)
    cell_label.font = font_doc_label
    cell_label.fill = fill_doc_label
    cell_label.alignment = align_center
    cell_label.border = thin_border

    ws.merge_cells(f'C{row}:E{row}')
    cell_content = ws.cell(row=row, column=3, value=value)
    cell_content.font = font_doc_value
    cell_content.alignment = align_left
    cell_content.border = thin_border
    for c in range(3, 6):
        ws.cell(row=row, column=c).border = thin_border

    ws.row_dimensions[row].height = 26
    return row + 1


def add_version_table(ws, row, versions):
    """
    版本记录表格（页签1）。
    参数: versions: [('日期', '版本', '修改内容', '负责人'), ...]
    """
    row = add_title_2(ws, row, '版本记录')
    headers = ['日期', '版本', '修改内容', '负责人']
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=i + 2, value=h)
        cell.font = font_table_header
        cell.fill = fill_table_header
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[row].height = 24
    row += 1
    for data in versions:
        for i, d in enumerate(data):
            cell = ws.cell(row=row, column=i + 2, value=d)
            cell.font = font_table_content
            cell.alignment = align_center
            cell.border = thin_border
        ws.row_dimensions[row].height = 22
        row += 1
    return row


def add_people_table(ws, row, people):
    """
    相关人员表格（页签1）。
    参数: people: [('角色', '姓名'), ...]
    """
    row = add_title_2(ws, row, '相关人员')
    headers = ['角色', '姓名']
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=i + 2, value=h)
        cell.font = font_table_header
        cell.fill = fill_table_header
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[row].height = 24
    row += 1
    for data in people:
        for i, d in enumerate(data):
            cell = ws.cell(row=row, column=i + 2, value=d)
            cell.font = font_table_content
            cell.alignment = align_center
            cell.border = thin_border
        ws.row_dimensions[row].height = 22
        row += 1
    return row


# ==================== 通用表格函数 ====================

def add_table_title(ws, row, text):
    """表格标题 - B列, 淡蓝底深蓝字, 12pt加粗"""
    cell = ws.cell(row=row, column=2, value=text)
    cell.font = Font(name='微软雅黑', size=12, bold=True, color=NAVY)
    cell.fill = fill_table_title
    cell.alignment = align_left
    ws.row_dimensions[row].height = 26
    return row + 1


def add_table(ws, row, title, headers, data):
    """
    一站式添加完整表格：标题 + 表头 + 数据行（交替行色）。

    参数:
        title: 表格标题, 如「表1-排位模式配置」
        headers: 表头列表, 如 ['字段名', '类型', '说明', '示例值']
        data: 二维列表, 如 [['mode_id', 'int', '模式ID', '1'], ...]
    
    列宽自适应: 当表格列数超过标准列(B~E)时,会自动调整超出列的宽度。
    """
    row = add_table_title(ws, row, title)

    # 列宽自适应：如果表格列数>4(超出B~E)，自动设置额外列的宽度
    num_cols = len(headers)
    if num_cols > 4:
        for extra in range(4, num_cols):
            col_letter = get_column_letter(extra + 2)  # 从F列开始
            # 计算该列最大内容宽度
            max_len = len(str(headers[extra]))
            for row_data in data:
                if extra < len(row_data):
                    max_len = max(max_len, len(str(row_data[extra])))
            # 设置列宽：每个字符约2.2个单位，最小15，最大40
            ws.column_dimensions[col_letter].width = max(15, min(40, max_len * 2.2 + 4))

    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=i + 2, value=h)
        cell.font = font_table_header
        cell.fill = fill_table_header
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[row].height = 24
    row += 1

    for idx, row_data in enumerate(data):
        for i, value in enumerate(row_data):
            cell = ws.cell(row=row, column=i + 2, value=value)
            cell.font = font_table_content
            cell.alignment = align_center
            cell.border = thin_border
            if idx % 2 == 1:
                cell.fill = fill_table_alt
        ws.row_dimensions[row].height = 22
        row += 1

    return row


# ==================== 图片嵌入函数 ====================

def add_image(ws, row, image_path, col=5, width_cm=16, description=None):
    """
    在指定位置嵌入图片。

    参数:
        ws: 工作表对象
        row: 起始行号
        image_path: 图片文件的绝对路径
        col: 图片锚点列号(默认E列=5)
        width_cm: 图片显示宽度(厘米), 默认16cm
        description: 图片下方的描述文字(可选, 灰色注释样式)

    返回:
        int: 下一行行号(图片占用的行数后)

    注意:
        - 图片会按比例缩放到指定宽度
        - 行高会自动调整以容纳图片
        - 支持 png/jpg/jpeg 格式
    """
    if not os.path.exists(image_path):
        # 图片不存在时退化为注释文本
        cell = ws.cell(row=row, column=col, value=f'[图片缺失：{os.path.basename(image_path)}]')
        cell.font = font_comment
        cell.alignment = align_left
        ws.row_dimensions[row].height = 22
        return row + 1

    img = XlImage(image_path)

    # 按宽度等比缩放
    target_width_px = width_cm * 37.8  # 1cm ≈ 37.8px
    scale = target_width_px / img.width
    img.width = target_width_px
    img.height = img.height * scale

    # 计算图片需要占用多少行(按每行22px估算)
    rows_needed = max(1, int(img.height / 22) + 1)

    # 设置锚点
    cell_ref = f'{get_column_letter(col)}{row}'
    ws.add_image(img, cell_ref)

    # 预留图片行高
    for r in range(row, row + rows_needed):
        ws.row_dimensions[r].height = 22

    next_row = row + rows_needed

    # 可选的图片描述
    if description:
        cell = ws.cell(row=next_row, column=col, value=description)
        cell.font = font_comment
        cell.alignment = align_left
        ws.row_dimensions[next_row].height = 20
        next_row += 1

    return next_row


def add_image_under_title(ws, row, image_path, width_cm=16, description=None):
    """
    在当前行的B列位置嵌入图片（用于标题下方的大图展示）。

    参数:
        ws: 工作表对象
        row: 起始行号
        image_path: 图片文件的绝对路径
        width_cm: 图片显示宽度(厘米), 默认16cm
        description: 图片下方的描述文字(可选)

    返回:
        int: 下一行行号
    """
    return add_image(ws, row, image_path, col=2, width_cm=width_cm, description=description)


def add_image_in_rule(ws, row, image_path, width_cm=14, description=None):
    """
    在规则的「2、交互图」位置嵌入图片（E列）。

    参数:
        ws: 工作表对象
        row: 起始行号(应在add_label('2、交互图')之后)
        image_path: 图片文件的绝对路径
        width_cm: 图片显示宽度(厘米), 默认14cm(E列内)
        description: 图片下方的描述文字(可选)

    返回:
        int: 下一行行号

    使用示例:
        row = add_label(ws, row, '2、交互图')
        row = add_image_in_rule(ws, row, 'path/to/flow_chart.png', description='图1：功能触发流程')
    """
    return add_image(ws, row, image_path, col=5, width_cm=width_cm, description=description)

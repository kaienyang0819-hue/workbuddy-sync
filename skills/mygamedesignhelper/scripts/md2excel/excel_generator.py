# -*- coding: utf-8 -*-
"""
Excel 生成器 - 根据 Document 数据模型生成标准格式的 .xlsx 文件

生成流程:
  1. 创建工作簿
  2. 生成"文档维护" Sheet (目录 + 元信息 + 修订记录)
  3. 按顺序生成各系统 Sheet（层级缩进: 功能B列/规则C列/细项D列）
  4. 生成"数值表格设计" Sheet（汇总所有表格完整版）
  5. 生成"数据打点及tlog" Sheet
  6. 生成"数据统计需求" Sheet (如有)
  7. 生成"经验和教训" Sheet (如有)
  8. 保存文件
"""
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont

from .models import Document, SystemSection, ContentBlock
from . import styles

# 尝试导入图片支持
try:
    from openpyxl.drawing.image import Image as XlImage
    HAS_IMAGE_SUPPORT = True
except ImportError:
    HAS_IMAGE_SUPPORT = False

# 数值表格设计 Sheet 名称常量
SHEET_DATA_TABLES = "数值表格设计"
SHEET_TLOG = "数据打点及tlog"


class ExcelGenerator:
    """Excel 策划文档生成器"""

    def __init__(self, doc: Document):
        self.doc = doc
        self.wb = Workbook()
        # 删除默认 Sheet
        self.wb.remove(self.wb.active)
        self._stats = {
            "total_cells": 0,
            "sheets_created": 0,
            "images_inserted": 0,
            "images_placeholder": 0,
            "images_not_found": 0,
        }
        # 收集所有系统Sheet中的表格，用于数值表格设计Sheet
        # 格式: [(system_name, table_name, ContentBlock), ...]
        self._collected_tables = []
        # 记录每个表格在"数值表格设计"Sheet中的行号，用于超链接
        self._table_anchors = {}  # table_name -> row_number

    def generate(self, output_path: str):
        """生成 Excel 文件"""
        # 先收集所有表格
        self._collected_tables = self.doc.collect_all_tables()

        # 预计算表格在"数值表格设计"Sheet中的锚点行号
        self._pre_calculate_table_anchors()

        # 1. 文档维护
        self._create_maintenance_sheet()

        # 2. 系统 Sheet
        for section in self.doc.systems:
            self._create_content_sheet(section, is_system=True)

        # 3. 数值表格设计（固定Sheet）
        self._create_data_tables_sheet()

        # 4. 数据打点及tlog（固定Sheet）
        self._create_tlog_sheet()

        # 5. 数据统计需求
        if self.doc.data_stats:
            self._create_content_sheet(self.doc.data_stats, is_system=False)

        # 6. 经验和教训
        if self.doc.lessons:
            self._create_content_sheet(self.doc.lessons, is_system=False)

        # 保存
        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(out))

        return self._stats

    # ============================================================
    #  文档维护 Sheet
    # ============================================================
    def _create_maintenance_sheet(self):
        ws = self.wb.create_sheet("文档维护")
        self._stats["sheets_created"] += 1

        # 列宽
        ws.column_dimensions["A"].width = styles.COL_A_WIDTH
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 3
        ws.column_dimensions["E"].width = styles.COL_E_WIDTH
        ws.column_dimensions["F"].width = styles.COL_F_WIDTH
        ws.column_dimensions["G"].width = 15
        for col in ["H", "I", "J", "K", "L", "M", "N"]:
            ws.column_dimensions[col].width = 8

        # --- 行1-2: Sheet 标题 ---
        title_text = f"{self.doc.project_title} 设计文档"
        ws.merge_cells("A1:N2")
        cell = ws["A1"]
        cell.value = title_text
        styles.apply_sheet_title(cell)
        ws.row_dimensions[1].height = styles.ROW_TITLE_HEIGHT
        ws.row_dimensions[2].height = 20
        self._stats["total_cells"] += 1

        # --- 行3: 空行 ---
        row = 4

        # --- 右侧: 元信息区域标题 ---
        ws.merge_cells(f"E{row}:N{row}")
        c = ws.cell(row=row, column=5, value="文档维护")
        styles.apply_heading2(c)
        for col_idx in range(6, 15):
            mc = ws.cell(row=row, column=col_idx)
            mc.fill = styles.FILL_MAROON

        # --- 左侧: 目录导航 (B列) ---
        all_sheets = self.doc.all_sheet_names
        nav_row = row

        for sheet_name in all_sheets:
            if sheet_name == "文档维护":
                continue
            c = ws.cell(row=nav_row, column=2, value=sheet_name)
            styles.apply_nav_link(c)
            c.hyperlink = f"#'{sheet_name}'!A1"
            self._stats["total_cells"] += 1
            nav_row += 1

        # --- 右侧: 元信息表格 ---
        meta = self.doc.meta
        meta_fields = [
            ("撰写人", meta.author),
            ("策划负责人", meta.planner),
            ("程序负责人", meta.programmer),
            ("美术负责人", meta.artist),
            ("创建日期", meta.created_date),
        ]

        meta_row = row + 1
        for label, value in meta_fields:
            ws.merge_cells(f"E{meta_row}:F{meta_row}")
            ws.merge_cells(f"G{meta_row}:N{meta_row}")

            lc = ws.cell(row=meta_row, column=5, value=label)
            styles.apply_meta_label(lc)
            fc = ws.cell(row=meta_row, column=6)
            fc.border = styles.BORDER_THIN_ALL

            vc = ws.cell(row=meta_row, column=7, value=value)
            styles.apply_meta_value(vc)
            for ci in range(8, 15):
                ws.cell(row=meta_row, column=ci).border = styles.BORDER_THIN_ALL

            self._stats["total_cells"] += 1
            meta_row += 1

        # 文档状态
        ws.merge_cells(f"E{meta_row}:F{meta_row}")
        ws.merge_cells(f"G{meta_row}:N{meta_row}")
        lc = ws.cell(row=meta_row, column=5, value="文档状态")
        styles.apply_meta_label(lc)
        ws.cell(row=meta_row, column=6).border = styles.BORDER_THIN_ALL

        status = meta.status or "草案"
        status_options = ["草案", "正式发布", "正在修改", "注销"]
        status_text = "；".join(
            f"[√] {s}" if s == status else f"[  ] {s}" for s in status_options
        )
        vc = ws.cell(row=meta_row, column=7, value=status_text)
        styles.apply_meta_value(vc)
        for ci in range(8, 15):
            ws.cell(row=meta_row, column=ci).border = styles.BORDER_THIN_ALL
        self._stats["total_cells"] += 1
        meta_row += 1

        # 空行
        meta_row += 1

        # 修订记录区域
        ws.merge_cells(f"E{meta_row}:N{meta_row}")
        c = ws.cell(row=meta_row, column=5, value="文档修订")
        styles.apply_heading2(c)
        for ci in range(6, 15):
            ws.cell(row=meta_row, column=ci).fill = styles.FILL_MAROON
        meta_row += 1

        # 修订记录表头
        ws.merge_cells(f"E{meta_row}:F{meta_row}")
        ws.merge_cells(f"G{meta_row}:H{meta_row}")
        ws.merge_cells(f"I{meta_row}:N{meta_row}")

        for col_idx, header in [(5, "修订日期"), (7, "修改人"), (9, "修改内容")]:
            c = ws.cell(row=meta_row, column=col_idx, value=header)
            styles.apply_meta_label(c)

        for ci in [6, 8, 10, 11, 12, 13, 14]:
            ws.cell(row=meta_row, column=ci).border = styles.BORDER_THIN_ALL
            ws.cell(row=meta_row, column=ci).fill = styles.FILL_LIGHT_GRAY
        meta_row += 1

        # 预留3行空白修订记录
        for _ in range(3):
            ws.merge_cells(f"E{meta_row}:F{meta_row}")
            ws.merge_cells(f"G{meta_row}:H{meta_row}")
            ws.merge_cells(f"I{meta_row}:N{meta_row}")
            for ci in range(5, 15):
                ws.cell(row=meta_row, column=ci).border = styles.BORDER_THIN_ALL
            meta_row += 1

    # ============================================================
    #  内容 Sheet (系统/数据统计/经验教训)
    # ============================================================
    def _create_content_sheet(self, section: SystemSection, is_system: bool = False):
        """
        创建内容Sheet

        参数:
          section: 章节数据
          is_system: 是否是系统Sheet（决定是否使用层级缩进和系统标题）
        """
        name = section.name[:31]
        ws = self.wb.create_sheet(name)
        self._stats["sheets_created"] += 1

        # 列宽
        ws.column_dimensions["A"].width = styles.COL_A_WIDTH
        ws.column_dimensions["B"].width = styles.COL_B_WIDTH
        ws.column_dimensions["C"].width = styles.COL_C_WIDTH
        ws.column_dimensions["D"].width = styles.COL_D_WIDTH
        ws.column_dimensions["E"].width = 30  # 用于表格溢出

        # --- 行1-2: Sheet 标题 ---
        ws.merge_cells("A1:H2")
        cell = ws["A1"]
        cell.value = section.name
        styles.apply_sheet_title(cell)
        ws.row_dimensions[1].height = styles.ROW_TITLE_HEIGHT
        ws.row_dimensions[2].height = 20
        self._stats["total_cells"] += 1

        # --- 内容区 ---
        row = 4

        if is_system:
            # 系统Sheet: 先写系统标题，再写层级缩进内容
            row = self._write_system_content(ws, row, section)
        else:
            # 非系统Sheet（数据统计需求/经验和教训）: 沿用扁平布局
            row = self._write_flat_content(ws, row, section)

    def _write_system_content(self, ws: Worksheet, row: int, section: SystemSection) -> int:
        """
        写入系统Sheet内容，使用层级缩进布局

        层级→列映射:
          系统标题(##) → B列（暗红底，【系统名】格式）
          功能标题(###) → B列（暗红底）
          功能下正文/列表 → B列
          规则标题(####) → C列（深灰底）
          规则下正文/列表 → C列
          规则细项标题(#####) → D列（加粗黑字）
          规则细项下正文/列表 → D列
        """
        # 先写系统标题行
        sys_title = f"【{section.name}】"
        c = ws.cell(row=row, column=2, value=sys_title)
        styles.apply_system_title(c)
        ws.row_dimensions[row].height = styles.ROW_SYSTEM_TITLE_HEIGHT
        # 给系统标题行涂满背景色（B到D列）
        for col_idx in range(3, 5):
            mc = ws.cell(row=row, column=col_idx)
            mc.fill = styles.FILL_MAROON
        self._stats["total_cells"] += 1
        row += 2  # 空一行

        # 当前列号跟踪（正文跟随上层标题所在列）
        current_col = 2  # 默认在B列（功能层级）

        for block in section.blocks:
            btype = block.type

            if btype == "heading3":
                # ### 功能 → B列（暗红底）
                current_col = 2
                c = ws.cell(row=row, column=2, value=block.content)
                styles.apply_heading2(c)
                ws.row_dimensions[row].height = styles.ROW_HEADING_HEIGHT
                self._stats["total_cells"] += 1
                row += 2  # 标题后空一行

            elif btype == "heading4":
                # #### 规则 → C列（深灰底）
                current_col = 3
                c = ws.cell(row=row, column=3, value=block.content)
                styles.apply_heading3(c)
                ws.row_dimensions[row].height = styles.ROW_HEADING_HEIGHT
                self._stats["total_cells"] += 1
                row += 2

            elif btype == "heading5":
                # ##### 规则细项 → D列（加粗黑字）
                current_col = 4
                c = ws.cell(row=row, column=4, value=block.content)
                styles.apply_heading5(c)
                self._stats["total_cells"] += 1
                row += 1

            elif btype == "table":
                # 表格：系统Sheet中只写简版(表头+最多3行)+超链接
                row = self._write_table_brief(ws, row, block, current_col)

            else:
                # 其他所有内容块（正文/列表/图片等）写到当前列
                row = self._write_block_at_col(ws, row, block, current_col)

        return row

    def _write_flat_content(self, ws: Worksheet, row: int, section: SystemSection) -> int:
        """
        写入扁平布局内容（用于非系统Sheet如数据统计需求、经验和教训）
        所有内容从B列开始，使用原始逻辑
        """
        for block in section.blocks:
            row = self._write_block_at_col(ws, row, block, col=2)
        return row

    def _write_block_at_col(self, ws: Worksheet, row: int, block: ContentBlock, col: int) -> int:
        """
        在指定列写入一个内容块
        返回: 下一个可用行号
        """
        btype = block.type

        if btype == "heading3":
            c = ws.cell(row=row, column=col, value=block.content)
            styles.apply_heading2(c)
            ws.row_dimensions[row].height = styles.ROW_HEADING_HEIGHT
            self._stats["total_cells"] += 1
            return row + 2

        elif btype == "heading4":
            c = ws.cell(row=row, column=col, value=block.content)
            styles.apply_heading3(c)
            ws.row_dimensions[row].height = styles.ROW_HEADING_HEIGHT
            self._stats["total_cells"] += 1
            return row + 2

        elif btype == "heading5":
            c = ws.cell(row=row, column=col, value=block.content)
            styles.apply_heading5(c)
            self._stats["total_cells"] += 1
            return row + 1

        elif btype == "text":
            c = ws.cell(row=row, column=col, value=block.content)
            styles.apply_body(c)
            self._stats["total_cells"] += 1
            return row + 1

        elif btype == "bold_text":
            c = ws.cell(row=row, column=col, value=block.content)
            styles.apply_emphasis(c)
            self._stats["total_cells"] += 1
            return row + 1

        elif btype == "mixed_text":
            row = self._write_mixed_text(ws, row, block, col)
            return row

        elif btype in ("list_item", "ordered_list_item"):
            prefix = "• "
            text = prefix + block.content
            if block.segments:
                row = self._write_mixed_list_item(ws, row, block, prefix, col)
            else:
                c = ws.cell(row=row, column=col, value=text)
                styles.apply_body(c)
                self._stats["total_cells"] += 1
            return row + 1

        elif btype == "sub_list_item":
            text = "    ◆ " + block.content
            if block.segments:
                row = self._write_mixed_list_item(ws, row, block, "    ◆ ", col)
            else:
                c = ws.cell(row=row, column=col, value=text)
                styles.apply_body(c)
                self._stats["total_cells"] += 1
            return row + 1

        elif btype == "image":
            row = self._write_image(ws, row, block, col)
            return row

        elif btype == "image_placeholder":
            text = f"[待插入图片: {block.image_desc}]"
            c = ws.cell(row=row, column=col, value=text)
            styles.apply_placeholder(c)
            self._stats["total_cells"] += 1
            self._stats["images_placeholder"] += 1
            return row + 2

        elif btype == "table":
            row = self._write_table_full(ws, row, block, col)
            return row

        else:
            # raw 或未知类型 → 作为正文写入（不丢内容）
            c = ws.cell(row=row, column=col, value=block.content)
            styles.apply_body(c)
            self._stats["total_cells"] += 1
            return row + 1

    # ============================================================
    #  混合格式文本
    # ============================================================
    def _write_mixed_text(self, ws: Worksheet, row: int, block: ContentBlock, col: int = 2) -> int:
        """写入混合格式文本（同一行含普通文字和加粗重点）"""
        try:
            parts = []
            for seg_type, seg_text in block.segments:
                if seg_type == "bold":
                    parts.append(TextBlock(
                        InlineFont(
                            rFont=styles.FONT_FAMILY,
                            sz=900,
                            color=styles.COLOR_RED,
                        ),
                        seg_text
                    ))
                else:
                    parts.append(TextBlock(
                        InlineFont(
                            rFont=styles.FONT_FAMILY,
                            sz=900,
                            color=styles.COLOR_BLACK,
                        ),
                        seg_text
                    ))

            rich_text = CellRichText(parts)
            c = ws.cell(row=row, column=col)
            c.value = rich_text
            c.alignment = styles.ALIGN_LEFT_CENTER
            self._stats["total_cells"] += 1
            return row + 1

        except Exception:
            fallback_text = block.content
            c = ws.cell(row=row, column=col, value=fallback_text)
            styles.apply_body(c)
            self._stats["total_cells"] += 1
            return row + 1

    def _write_mixed_list_item(self, ws: Worksheet, row: int, block: ContentBlock, prefix: str, col: int = 2) -> int:
        """写入含混合格式的列表项"""
        try:
            parts = [TextBlock(
                InlineFont(rFont=styles.FONT_FAMILY, sz=900, color=styles.COLOR_BLACK),
                prefix
            )]
            for seg_type, seg_text in block.segments:
                if seg_type == "bold":
                    parts.append(TextBlock(
                        InlineFont(rFont=styles.FONT_FAMILY, sz=900, color=styles.COLOR_RED),
                        seg_text
                    ))
                else:
                    parts.append(TextBlock(
                        InlineFont(rFont=styles.FONT_FAMILY, sz=900, color=styles.COLOR_BLACK),
                        seg_text
                    ))

            rich_text = CellRichText(parts)
            c = ws.cell(row=row, column=col)
            c.value = rich_text
            c.alignment = styles.ALIGN_LEFT_CENTER
            self._stats["total_cells"] += 1
            return row

        except Exception:
            text = prefix + block.content
            c = ws.cell(row=row, column=col, value=text)
            styles.apply_body(c)
            self._stats["total_cells"] += 1
            return row

    # ============================================================
    #  图片
    # ============================================================
    def _write_image(self, ws: Worksheet, row: int, block: ContentBlock, col: int = 2) -> int:
        """插入图片或写入占位文字"""
        img_path = block.image_path

        if HAS_IMAGE_SUPPORT and img_path and os.path.isfile(img_path):
            try:
                img = XlImage(img_path)
                max_width = 600
                max_height = 400
                if img.width > max_width:
                    ratio = max_width / img.width
                    img.width = max_width
                    img.height = int(img.height * ratio)
                if img.height > max_height:
                    ratio = max_height / img.height
                    img.height = max_height
                    img.width = int(img.width * ratio)

                cell_ref = f"{get_column_letter(col)}{row}"
                ws.add_image(img, cell_ref)
                self._stats["images_inserted"] += 1

                if block.image_desc:
                    desc_row = row + int(img.height / 20) + 1
                    c = ws.cell(row=desc_row, column=col, value=f"图: {block.image_desc}")
                    styles.apply_body(c)
                    self._stats["total_cells"] += 1
                    return desc_row + 2
                return row + int(img.height / 20) + 2

            except Exception:
                pass

        if img_path:
            text = f"[图片未找到: {img_path}]"
        else:
            text = f"[待插入图片: {block.image_desc}]"
        c = ws.cell(row=row, column=col, value=text)
        styles.apply_placeholder(c)
        self._stats["total_cells"] += 1
        self._stats["images_not_found"] += 1
        return row + 2

    # ============================================================
    #  表格处理
    # ============================================================
    def _write_table_full(self, ws: Worksheet, row: int, block: ContentBlock, col: int = 2) -> int:
        """写入完整表格（用于数值表格设计Sheet和非系统Sheet）"""
        start_col = col

        # 表头
        if block.table_headers:
            for ci, header in enumerate(block.table_headers):
                c = ws.cell(row=row, column=start_col + ci, value=header)
                styles.apply_table_header(c)
                self._stats["total_cells"] += 1
            row += 1

        # 数据行
        for table_row in block.table_rows:
            for ci, cell_val in enumerate(table_row):
                c = ws.cell(row=row, column=start_col + ci, value=cell_val)
                styles.apply_table_data(c)
                self._stats["total_cells"] += 1
            row += 1

        return row + 1  # 表格后空一行

    def _write_table_brief(self, ws: Worksheet, row: int, block: ContentBlock, col: int = 2) -> int:
        """
        写入简版表格（用于系统Sheet）
        只保留表头 + 最多3行示意数据 + 超链接到数值表格设计Sheet
        """
        start_col = col
        table_name = block.table_name

        # 表头
        if block.table_headers:
            for ci, header in enumerate(block.table_headers):
                c = ws.cell(row=row, column=start_col + ci, value=header)
                styles.apply_table_header(c)
                self._stats["total_cells"] += 1

            # 在表头行最后一列旁边加超链接
            link_col = start_col + len(block.table_headers)
            anchor_row = self._table_anchors.get(table_name, 1)
            link_cell = ws.cell(row=row, column=link_col, value="→ 查看完整表格")
            styles.apply_table_link(link_cell)
            link_cell.hyperlink = f"#'{SHEET_DATA_TABLES}'!B{anchor_row}"
            self._stats["total_cells"] += 1
            row += 1

        # 最多显示3行数据
        max_brief_rows = min(3, len(block.table_rows))
        for ri in range(max_brief_rows):
            table_row = block.table_rows[ri]
            for ci, cell_val in enumerate(table_row):
                c = ws.cell(row=row, column=start_col + ci, value=cell_val)
                styles.apply_table_data(c)
                self._stats["total_cells"] += 1
            row += 1

        # 如果有更多行，写一个省略提示
        if len(block.table_rows) > max_brief_rows:
            msg = f"...（共{len(block.table_rows)}行，完整数据见'数值表格设计'Sheet）"
            c = ws.cell(row=row, column=start_col, value=msg)
            styles.apply_placeholder(c)
            self._stats["total_cells"] += 1
            row += 1

        return row + 1  # 表格后空一行

    # ============================================================
    #  数值表格设计 Sheet（固定）
    # ============================================================
    def _create_data_tables_sheet(self):
        """创建"数值表格设计"Sheet，汇总所有系统Sheet中的表格完整版"""
        ws = self.wb.create_sheet(SHEET_DATA_TABLES)
        self._stats["sheets_created"] += 1

        # 列宽
        ws.column_dimensions["A"].width = styles.COL_A_WIDTH
        ws.column_dimensions["B"].width = 20
        for col_letter in ["C", "D", "E", "F", "G", "H", "I", "J"]:
            ws.column_dimensions[col_letter].width = 18

        # --- 行1-2: Sheet 标题 ---
        ws.merge_cells("A1:J2")
        cell = ws["A1"]
        cell.value = "数值表格设计"
        styles.apply_sheet_title(cell)
        ws.row_dimensions[1].height = styles.ROW_TITLE_HEIGHT
        ws.row_dimensions[2].height = 20
        self._stats["total_cells"] += 1

        # --- 区域标题 ---
        row = 4
        section_title = f"【数值表格设计】"
        c = ws.cell(row=row, column=2, value=section_title)
        styles.apply_system_title(c)
        ws.row_dimensions[row].height = styles.ROW_SYSTEM_TITLE_HEIGHT
        for col_idx in range(3, 11):
            ws.cell(row=row, column=col_idx).fill = styles.FILL_MAROON
        self._stats["total_cells"] += 1
        row += 2

        if not self._collected_tables:
            c = ws.cell(row=row, column=2, value="（暂无表格数据）")
            styles.apply_placeholder(c)
            self._stats["total_cells"] += 1
            return

        # 先预计算每个表格的锚点行号（第一遍，用于系统Sheet中的超链接）
        # 由于此方法在系统Sheet之后调用，需要先计算再回填
        # 实际做法：先写入数值表格设计Sheet，记录行号，然后系统Sheet中已经用了
        # 所以我们需要在 generate() 开始时预先计算
        # —— 已通过 _pre_calculate_table_anchors() 在 generate() 中调用

        for idx, (system_name, table_name, table_block) in enumerate(self._collected_tables):
            # 记录锚点
            self._table_anchors[table_name] = row

            # 表名标题行
            c = ws.cell(row=row, column=2, value=f"{table_name}（来源: {system_name}）")
            styles.apply_table_name(c)
            self._stats["total_cells"] += 1
            row += 1

            # 完整表格
            row = self._write_table_full(ws, row, table_block, col=2)
            row += 1  # 额外空一行分隔

    # ============================================================
    #  数据打点及tlog Sheet（固定）
    # ============================================================
    def _create_tlog_sheet(self):
        """创建"数据打点及tlog"Sheet"""
        ws = self.wb.create_sheet(SHEET_TLOG)
        self._stats["sheets_created"] += 1

        # 列宽
        ws.column_dimensions["A"].width = styles.COL_A_WIDTH
        ws.column_dimensions["B"].width = 8   # 序号
        ws.column_dimensions["C"].width = 18  # 所属模块
        ws.column_dimensions["D"].width = 35  # 事件描述
        ws.column_dimensions["E"].width = 10  # 优先级
        ws.column_dimensions["F"].width = 25  # 接口ID/参数
        ws.column_dimensions["G"].width = 12  # 负责人

        # --- 行1-2: Sheet 标题 ---
        ws.merge_cells("A1:G2")
        cell = ws["A1"]
        cell.value = "数据打点及tlog"
        styles.apply_sheet_title(cell)
        ws.row_dimensions[1].height = styles.ROW_TITLE_HEIGHT
        ws.row_dimensions[2].height = 20
        self._stats["total_cells"] += 1

        # --- 区域标题 ---
        row = 4
        section_title = "【tlog及打点设计】"
        c = ws.cell(row=row, column=2, value=section_title)
        styles.apply_system_title(c)
        ws.row_dimensions[row].height = styles.ROW_SYSTEM_TITLE_HEIGHT
        for col_idx in range(3, 8):
            ws.cell(row=row, column=col_idx).fill = styles.FILL_MAROON
        self._stats["total_cells"] += 1
        row += 2

        # 如果MD中有tlog章节，写入内容
        if self.doc.tlog and self.doc.tlog.blocks:
            # 尝试识别tlog章节中的表格
            has_table = False
            for block in self.doc.tlog.blocks:
                if block.type == "table":
                    has_table = True
                    break

            if has_table:
                # 有表格数据，按表格格式写入
                for block in self.doc.tlog.blocks:
                    if block.type == "table":
                        row = self._write_table_full(ws, row, block, col=2)
                    elif block.type in ("heading3", "heading4", "heading5"):
                        c = ws.cell(row=row, column=2, value=block.content)
                        styles.apply_table_name(c)
                        self._stats["total_cells"] += 1
                        row += 1
                    elif block.type in ("text", "bold_text", "mixed_text", "list_item",
                                        "ordered_list_item", "sub_list_item", "raw"):
                        row = self._write_block_at_col(ws, row, block, col=2)
            else:
                # 无表格数据，按普通内容写入
                for block in self.doc.tlog.blocks:
                    row = self._write_block_at_col(ws, row, block, col=2)
        else:
            # 没有tlog章节，写一个空白模板
            row = self._write_tlog_template(ws, row)

    def _write_tlog_template(self, ws: Worksheet, row: int) -> int:
        """写入空白的tlog模板表格"""
        # 表名
        c = ws.cell(row=row, column=2, value="表1-事件打点列表")
        styles.apply_table_name(c)
        self._stats["total_cells"] += 1
        row += 1

        # 表头
        headers = ["序号", "所属模块", "事件描述", "优先级", "接口ID/参数", "负责人"]
        for ci, header in enumerate(headers):
            c = ws.cell(row=row, column=2 + ci, value=header)
            styles.apply_table_header(c)
            self._stats["total_cells"] += 1
        row += 1

        # 预留5行空白数据行
        for ri in range(1, 6):
            ws.cell(row=row, column=2, value=str(ri))
            for ci in range(6):
                c = ws.cell(row=row, column=2 + ci)
                styles.apply_table_data(c)
                if ci == 0:
                    c.value = str(ri)
                self._stats["total_cells"] += 1
            row += 1

        return row + 1

    def _pre_calculate_table_anchors(self):
        """
        预计算每个表格在"数值表格设计"Sheet中的起始行号
        这样系统Sheet中的超链接可以正确指向
        """
        row = 6  # 数值表格设计Sheet: 行1-2标题, 行4系统标题, 行5空, 从行6开始

        for idx, (system_name, table_name, table_block) in enumerate(self._collected_tables):
            self._table_anchors[table_name] = row
            row += 1  # 表名标题行

            # 表头行
            if table_block.table_headers:
                row += 1
            # 数据行
            row += len(table_block.table_rows)
            row += 2  # 表格后空行 + 额外空行


def generate_excel(doc: Document, output_path: str) -> dict:
    """
    便捷函数: 从 Document 生成 Excel 文件

    参数:
      doc: Document 数据模型
      output_path: 输出 .xlsx 文件路径

    返回:
      统计信息 dict
    """
    gen = ExcelGenerator(doc)
    return gen.generate(output_path)

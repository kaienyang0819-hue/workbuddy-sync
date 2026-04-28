# -*- coding: utf-8 -*-
"""
Excel 样式配置 - 集中定义所有格式常量

所有颜色、字体、对齐、边框样式都在此文件中定义，
对应"空白模版参考_系统&玩法类策划文档.xlsx"模板的格式系统。
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from copy import copy

# ============================================================
#  颜色常量 (RGB hex, 不含 FF 前缀)
# ============================================================
COLOR_DARK_GRAY = "404040"       # 深灰底 (模板 theme=1, tint=0.25 近似)
COLOR_MAROON = "800000"          # 暗红底 (模板 FF800000)
COLOR_WHITE = "FFFFFF"           # 白色字
COLOR_BLACK = "000000"           # 黑色正文
COLOR_RED = "FF0000"             # 重点内容（加粗标注）
COLOR_LIGHT_GRAY = "F2F2F2"     # 浅灰底（元信息表格区）
COLOR_PLACEHOLDER = "999999"     # 图片占位符灰色
COLOR_LINK = "0563C1"            # 超链接蓝色

# ============================================================
#  字体
# ============================================================
FONT_FAMILY = "微软雅黑"

# Sheet 标题 (行1-2, 大标题)
FONT_SHEET_TITLE = Font(
    name=FONT_FAMILY, size=20, bold=True, color=COLOR_WHITE
)

# 系统标题 (B列系统名称，如 【宠物获取与孵化】)
FONT_SYSTEM_TITLE = Font(
    name=FONT_FAMILY, size=14, bold=True, color=COLOR_WHITE
)

# 2级标题 (### 功能, 暗红底)
FONT_HEADING2 = Font(
    name=FONT_FAMILY, size=11, bold=True, color=COLOR_WHITE
)

# 3级标题 (#### 规则, 深灰底)
FONT_HEADING3 = Font(
    name=FONT_FAMILY, size=11, bold=True, color=COLOR_WHITE
)

# 5级小标题 (##### 规则细项)
FONT_HEADING5 = Font(
    name=FONT_FAMILY, size=11, bold=True, color=COLOR_BLACK
)

# 正文
FONT_BODY = Font(
    name=FONT_FAMILY, size=9, color=COLOR_BLACK
)

# 重点内容 (MD 中 **加粗**)
FONT_EMPHASIS = Font(
    name=FONT_FAMILY, size=9, color=COLOR_RED
)

# 图片占位符
FONT_PLACEHOLDER = Font(
    name=FONT_FAMILY, size=9, color=COLOR_PLACEHOLDER, italic=True
)

# 元信息标签
FONT_META_LABEL = Font(
    name=FONT_FAMILY, size=11, color=COLOR_BLACK
)

# 元信息值
FONT_META_VALUE = Font(
    name=FONT_FAMILY, size=11, color=COLOR_BLACK
)

# 目录导航 (超链接)
FONT_NAV_LINK = Font(
    name=FONT_FAMILY, size=11, color=COLOR_LINK, underline="single"
)

# 目录分类标题 (加粗深灰底)
FONT_NAV_CATEGORY = Font(
    name=FONT_FAMILY, size=11, bold=True, color=COLOR_WHITE
)

# 表格表头（系统Sheet内表格 + 固定Sheet表格）
FONT_TABLE_HEADER = Font(
    name=FONT_FAMILY, size=9, bold=True, color=COLOR_WHITE
)

# 表格数据
FONT_TABLE_DATA = Font(
    name=FONT_FAMILY, size=9, color=COLOR_BLACK
)

# 表名标题 (如 "表1-xxx配置表"，加粗黑字)
FONT_TABLE_NAME = Font(
    name=FONT_FAMILY, size=11, bold=True, color=COLOR_BLACK
)

# 超链接 (系统Sheet中表格跳转链接)
FONT_TABLE_LINK = Font(
    name=FONT_FAMILY, size=9, color=COLOR_LINK, underline="single"
)

# ============================================================
#  填充 (背景色)
# ============================================================
FILL_DARK_GRAY = PatternFill(
    patternType="solid", fgColor=COLOR_DARK_GRAY
)

FILL_MAROON = PatternFill(
    patternType="solid", fgColor=COLOR_MAROON
)

FILL_LIGHT_GRAY = PatternFill(
    patternType="solid", fgColor=COLOR_LIGHT_GRAY
)

FILL_NONE = PatternFill(fill_type=None)

# ============================================================
#  对齐
# ============================================================
ALIGN_LEFT_CENTER = Alignment(
    horizontal="left", vertical="center", wrap_text=True
)

ALIGN_CENTER_CENTER = Alignment(
    horizontal="center", vertical="center", wrap_text=True
)

ALIGN_LEFT_TOP = Alignment(
    horizontal="left", vertical="top", wrap_text=True
)

# ============================================================
#  边框
# ============================================================
THIN_SIDE = Side(style="thin", color=COLOR_BLACK)

BORDER_THIN_ALL = Border(
    left=THIN_SIDE, right=THIN_SIDE,
    top=THIN_SIDE, bottom=THIN_SIDE
)

BORDER_NONE = Border()

# ============================================================
#  列宽和行高
# ============================================================
COL_A_WIDTH = 5.33       # A列极窄
COL_B_WIDTH = 50         # B列（系统标题+功能标题）
COL_C_WIDTH = 50         # C列（规则标题+正文）
COL_D_WIDTH = 50         # D列（规则细项+正文）
COL_E_WIDTH = 12         # E列(元信息标签)
COL_F_WIDTH = 12         # F列(元信息标签续)
COL_G_WIDTH = 15         # G列(元信息值)

ROW_TITLE_HEIGHT = 40    # 标题行高度
ROW_HEADING_HEIGHT = 25  # 标题行高度
ROW_SYSTEM_TITLE_HEIGHT = 30  # 系统标题行高度
ROW_BODY_HEIGHT = None   # 正文行高度(自动)

# ============================================================
#  组合样式 (便捷函数)
# ============================================================

def apply_sheet_title(cell):
    """应用 Sheet 标题样式 (行1-2)"""
    cell.font = FONT_SHEET_TITLE
    cell.fill = FILL_DARK_GRAY
    cell.alignment = ALIGN_LEFT_CENTER

def apply_system_title(cell):
    """应用系统标题样式 (B列系统名称 【xxx】)"""
    cell.font = FONT_SYSTEM_TITLE
    cell.fill = FILL_MAROON
    cell.alignment = ALIGN_LEFT_CENTER

def apply_heading2(cell):
    """应用2级标题样式 (### 功能, 暗红底)"""
    cell.font = FONT_HEADING2
    cell.fill = FILL_MAROON
    cell.alignment = ALIGN_LEFT_CENTER

def apply_heading3(cell):
    """应用3级标题样式 (#### 规则, 深灰底)"""
    cell.font = FONT_HEADING3
    cell.fill = FILL_DARK_GRAY
    cell.alignment = ALIGN_LEFT_CENTER

def apply_heading5(cell):
    """应用5级小标题样式 (##### 细项)"""
    cell.font = FONT_HEADING5
    cell.fill = FILL_NONE
    cell.alignment = ALIGN_LEFT_CENTER

def apply_body(cell):
    """应用正文样式"""
    cell.font = FONT_BODY
    cell.fill = FILL_NONE
    cell.alignment = ALIGN_LEFT_CENTER

def apply_emphasis(cell):
    """应用重点内容样式 (红色)"""
    cell.font = FONT_EMPHASIS
    cell.fill = FILL_NONE
    cell.alignment = ALIGN_LEFT_CENTER

def apply_placeholder(cell):
    """应用图片占位符样式 (灰色斜体)"""
    cell.font = FONT_PLACEHOLDER
    cell.fill = FILL_NONE
    cell.alignment = ALIGN_LEFT_CENTER

def apply_meta_label(cell):
    """应用元信息标签样式"""
    cell.font = FONT_META_LABEL
    cell.fill = FILL_LIGHT_GRAY
    cell.alignment = ALIGN_CENTER_CENTER
    cell.border = BORDER_THIN_ALL

def apply_meta_value(cell):
    """应用元信息值样式"""
    cell.font = FONT_META_VALUE
    cell.fill = FILL_NONE
    cell.alignment = ALIGN_CENTER_CENTER
    cell.border = BORDER_THIN_ALL

def apply_nav_link(cell):
    """应用目录导航链接样式"""
    cell.font = FONT_NAV_LINK
    cell.fill = FILL_NONE
    cell.alignment = ALIGN_LEFT_CENTER

def apply_nav_category(cell):
    """应用目录分类标题样式"""
    cell.font = FONT_NAV_CATEGORY
    cell.fill = FILL_DARK_GRAY
    cell.alignment = ALIGN_LEFT_CENTER

def apply_table_header(cell):
    """应用表格表头样式 (暗红/深灰底 + 白字 + 加粗)"""
    cell.font = FONT_TABLE_HEADER
    cell.fill = FILL_DARK_GRAY
    cell.alignment = ALIGN_LEFT_CENTER
    cell.border = BORDER_THIN_ALL

def apply_table_data(cell):
    """应用表格数据样式"""
    cell.font = FONT_TABLE_DATA
    cell.fill = FILL_NONE
    cell.alignment = ALIGN_LEFT_CENTER
    cell.border = BORDER_THIN_ALL

def apply_table_name(cell):
    """应用表名标题样式 (如 '表1-xxx配置表')"""
    cell.font = FONT_TABLE_NAME
    cell.fill = FILL_NONE
    cell.alignment = ALIGN_LEFT_CENTER

def apply_table_link(cell):
    """应用表格超链接样式"""
    cell.font = FONT_TABLE_LINK
    cell.fill = FILL_NONE
    cell.alignment = ALIGN_LEFT_CENTER
